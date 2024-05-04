from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import time

chrome_driver_path = "./chromedriver-win64/chromedriver.exe"
service = Service(executable_path=chrome_driver_path)
driver = webdriver.Chrome(service=service)

driver.maximize_window()
driver.get('https://news.walla.co.il/')
time.sleep(2)

articles = driver.find_elements(By.CSS_SELECTOR, "section.css-71im40 ul li a[href]")
data = []

for article in articles:
    main_window = driver.current_window_handle
    link = article.get_attribute('href')

    # פתיחת כרטיסייה חדשה ומעבר אליה
    driver.execute_script("window.open(arguments[0]);", link)
    time.sleep(2)
    new_window = [window for window in driver.window_handles if window != main_window][0]
    driver.switch_to.window(new_window)
    
    try:
        # כותרת
        title = driver.find_element(By.CSS_SELECTOR, 'h1.title.article_speakable').text

        # תקציר
        summary = driver.find_element(By.CSS_SELECTOR, 'p.subtitle.article_speakable').text

        # עדכון
        update_time = driver.find_element(By.CSS_SELECTOR, 'p.date-and-time-p').text

        # כתב
        author = driver.find_element(By.CSS_SELECTOR, 'div.writer-name-item > p').text

        # לינק לתמונה
        image_url = driver.find_element(By.CSS_SELECTOR, 'picture.ratio-desktop img').get_attribute('srcset').split()[0]

        # תוכן
        speakable_paragraphs = driver.find_elements(By.CSS_SELECTOR, "p.article_speakable")
        content = ""
        if len(speakable_paragraphs) > 2:  # בדיקה שיש לפח3 ;;;;;;;;ות שניים לאחר הראשון
            content = speakable_paragraphs[1].text + " " + speakable_paragraphs[2].text

        data.append([title, summary, update_time, author, image_url, content])

    except Exception as e:
        print(f"Error extracting data: {e}")
        content = "תקלה באיסוף נתונים"
    
    
    # סגירת הכרטיסייה הנוכחית וחזרה לראשית
    driver.close()
    driver.switch_to.window(main_window)
    
driver.quit()

############

# יצירת DataFrame
df = pd.DataFrame(data, columns=['כותרת', 'תקציר', 'עדכון', 'כתב', 'לינק לתמונה','תוכן'])
df.to_csv('wlla_mission.csv', index=False, encoding='utf-8-sig') # הכנת קובץ csv כמו במשימה
df.to_excel('walla_news.xlsx', index=False, engine='openpyxl') # הכנת קובץ excel לפני עיצוב

# טעינת ה-Workbook שנוצר
wb = load_workbook('walla_news.xlsx')
ws = wb.active

# הגדרות עיצוב
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
cell_font = Font(size=11)
cell_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
cell_alignment = Alignment(horizontal="right")

# עיצוב הכותרת
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center")

# עיצוב שאר הטבלה
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.font = cell_font
        cell.fill = cell_fill
        cell.border = thin_border
        cell.alignment = cell_alignment

# שמירת השינויים לקובץ
wb.save("walla_news.xlsx")